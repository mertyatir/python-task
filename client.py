import argparse
import pandas as pd
import requests
from datetime import datetime
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def str2bool(v):
    if isinstance(v, bool):
       return v
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')

# Parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('-k', '--keys', nargs='*', default=[])
parser.add_argument('-c', '--colored', type=str2bool, default=True)
args = parser.parse_args()

# Read CSV file
df = pd.read_csv('vehicles.csv', delimiter=';')

# Send POST request to server
response = requests.post('http://localhost:5000/api', data=df.to_csv(index=False).encode('utf-8'))

# Check if the request was successful
if response.status_code == 200:
    # Convert response to DataFrame
    df = pd.read_json(StringIO(response.text))
else:
    print(f"Error: Server returned status code {response.status_code}")
    print(response.text)

# Filter columns
if "colorCodes" in df.columns:
    columns = ['rnr', 'hu','colorCodes'] + args.keys
else:
    columns = ['rnr', 'hu'] + args.keys
df = df[columns]

# Check if 'hu' column exists
if 'hu' in df.columns:
    # Convert 'hu' column to datetime
    df['hu'] = pd.to_datetime(df['hu'])
    # Calculate the difference in months between 'hu' and the current date
    df['hu_diff'] = (df['hu'].dt.year - pd.Timestamp.now().year) * 12 + df['hu'].dt.month - pd.Timestamp.now().month
else:
    print("Error: 'hu' column not found in DataFrame")

# Sort DataFrame by 'gruppe' and reset index
df.sort_values('gruppe', inplace=True)
df.reset_index(drop=True, inplace=True)

# Write to Excel file
filename = f'vehicles_{datetime.now().isoformat()}.xlsx'
df.to_excel(filename, index=False)

# Open the Excel file with openpyxl
wb = load_workbook(filename)
ws = wb.active

# Apply color formatting
if args.colored:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        hu_diff = df.loc[row[0].row - 2, "hu_diff"]
        if hu_diff >= -3:
            fill = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
        elif hu_diff  >= -12:
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        else:
            fill = PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")
        for cell in row:
            cell.fill = fill
            

# Check if 'labelIds' and 'colorCodes' are in DataFrame columns and keys argument
if 'labelIds' in df.columns and 'colorCodes' in df.columns and 'labelIds' in args.keys:
    # Iterate over each row in DataFrame
    for index, row in df.iterrows():
        # Check if 'colorCodes' is not null
        if pd.notnull(row['colorCodes']).any():
            # Convert colorCodes to ARGB hex value
            color_codes = [code.lstrip('#') for code in row['colorCodes']]
            for color_code in color_codes:
                if len(color_code) == 6:
                    color_code = 'FF' + color_code
                # Use colorCode to tint the cell's text in Excel file
                ws.cell(row=index+2, column=columns.index('labelIds')+1).font = Font(color=color_code)
                break

# Get all column headers
headers = [cell.value for cell in ws[1]]

# Determine columns to keep
columns_to_keep = ['rnr'] + args.keys

# Delete columns not in columns_to_keep
for header in headers:
    if header not in columns_to_keep:
        # Find the index of the column
        col_index = headers.index(header) + 1  # 1-based index
        # Delete the column
        ws.delete_cols(col_index)

# Save the changes
wb.save(filename)




